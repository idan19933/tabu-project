"""Excel report generation for simulation results using openpyxl.

Two report types matching the Shikun & Binui client format:
  - Management Report (דוח ניהולי): Sections 1-3 (summary, state, program)
  - Economic Report (דוח כלכלי): Sections 4-5 (costs, revenue, profit)
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.simulation import Simulation

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
SECTION_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SECTION_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=10)
CELL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
CURRENCY_FMT = '#,##0'
DECIMAL_FMT = '#,##0.00'
PERCENT_FMT = '0.00"%"'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
RTL_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

GREEN_FONT = Font(name="Arial", size=10, color="006600", bold=True)
RED_FONT = Font(name="Arial", size=10, color="CC0000", bold=True)
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _safe(val, default=0):
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _style_row(ws, row: int, cols: int, fill=None, font=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        else:
            cell.font = CELL_FONT
        cell.alignment = RTL_ALIGN
        cell.border = THIN_BORDER


def _write_section_header(ws, row: int, title: str, cols: int = 4) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row=row, column=1, value=title)
    _style_row(ws, row, cols, fill=SECTION_FILL, font=SECTION_FONT)
    return row + 1


def _write_kv_row(ws, row: int, label: str, value, fmt: str | None = None,
                  bold: bool = False, highlight: str | None = None) -> int:
    """Write a label-value pair row."""
    ws.cell(row=row, column=1, value=label)
    cell_v = ws.cell(row=row, column=2, value=value)
    if fmt:
        cell_v.number_format = fmt
    font = BOLD_FONT if bold else CELL_FONT
    ws.cell(row=row, column=1).font = font
    cell_v.font = font
    for c in range(1, 3):
        ws.cell(row=row, column=c).alignment = RTL_ALIGN
        ws.cell(row=row, column=c).border = THIN_BORDER
    if highlight == "green":
        for c in range(1, 3):
            ws.cell(row=row, column=c).fill = LIGHT_GREEN_FILL
    elif highlight == "red":
        for c in range(1, 3):
            ws.cell(row=row, column=c).fill = LIGHT_RED_FILL
    return row + 1


def _write_table(ws, start_row: int, headers: list[str], rows: list[list],
                 col_formats: list[str | None] | None = None) -> int:
    cols = len(headers)
    for i, h in enumerate(headers):
        ws.cell(row=start_row, column=i + 1, value=h)
    _style_row(ws, start_row, cols, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)

    for r_idx, row_data in enumerate(rows):
        row_num = start_row + 1 + r_idx
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=row_num, column=c_idx + 1, value=val)
            if col_formats and c_idx < len(col_formats) and col_formats[c_idx]:
                cell.number_format = col_formats[c_idx]
        _style_row(ws, row_num, cols)

    return start_row + 1 + len(rows) + 1  # extra blank row


def _get_all_results(sim: Simulation) -> dict:
    """Extract all result values safely."""
    r = sim.simulation_results
    if not r:
        return {}
    from app.models.simulation_result import SimulationResult
    result = {}
    for col in SimulationResult.__table__.columns:
        if col.name == "simulation_id":
            continue
        val = getattr(r, col.name, None)
        if val is not None:
            try:
                result[col.name] = float(val)
            except (TypeError, ValueError):
                result[col.name] = val
        else:
            result[col.name] = 0 if col.name not in ("monthly_cash_flows", "calculation_details", "cost_breakdown", "revenue_breakdown", "area_breakdown") else ([] if col.name == "monthly_cash_flows" else {})
    return result


# ---------------------------------------------------------------------------
# Management Report (דוח ניהולי) — Sections 1-3
# ---------------------------------------------------------------------------

def generate_management_report(sim: Simulation) -> io.BytesIO:
    """Generate management summary Excel report.

    Sections:
      1. Project info + KPIs
      2. מצב יוצא (proposed state)
      3. פרוגרמה (building program)
      + Apartment mix table
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "דוח ניהולי"
    ws.sheet_view.rightToLeft = True

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    vals = _get_all_results(sim)
    project = sim.project if hasattr(sim, "project") else None
    tabu = (project.tabu_data or {}) if project else {}

    # === Title ===
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    title_cell = ws.cell(row=row, column=1, value="דוח ניהולי - בדיקת כדאיות כלכלית")
    title_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    row += 2

    # === Project Info ===
    row = _write_kv_row(ws, row, "פרויקט", project.name if project else sim.version_name)
    row = _write_kv_row(ws, row, "תאריך הפקה", datetime.now().strftime("%d/%m/%Y"))
    row = _write_kv_row(ws, row, "גרסת סימולציה", sim.version_name)
    # Planning stage (audit gap)
    pp = sim.planning_parameters
    if pp and getattr(pp, "planning_stage", None):
        row = _write_kv_row(ws, row, "שלב תכנוני", pp.planning_stage)
    if tabu.get("address"):
        row = _write_kv_row(ws, row, "כתובת", str(tabu["address"]))
    if tabu.get("block"):
        row = _write_kv_row(ws, row, "גוש/חלקה", f"{tabu.get('block', '')}/{tabu.get('parcel', '')}")
    if tabu.get("area_sqm"):
        row = _write_kv_row(ws, row, 'שטח מגרש (מ"ר)', _safe(tabu["area_sqm"]), DECIMAL_FMT)
    row += 1

    # === KPI Summary ===
    row = _write_section_header(ws, row, "מדדים עיקריים")
    profit = vals.get("expected_profit", vals.get("profit", 0))
    profit_hl = "green" if profit > 0 else "red"
    row = _write_kv_row(ws, row, "רווח יזמי (₪)", profit, CURRENCY_FMT, bold=True, highlight=profit_hl)
    row = _write_kv_row(ws, row, "שיעור רווחיות (%)", vals.get("profit_percent", vals.get("profitability_rate", 0)), DECIMAL_FMT, bold=True)
    row = _write_kv_row(ws, row, "רווחיות סטנדרט 21% (%)", vals.get("profit_percent_standard21", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "IRR (%)", vals.get("irr", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "NPV (₪)", vals.get("npv", 0), CURRENCY_FMT)
    row = _write_kv_row(ws, row, 'סה"כ הכנסות (₪)', vals.get("total_revenue", 0), CURRENCY_FMT)
    row = _write_kv_row(ws, row, 'סה"כ עלויות כולל מע"מ (₪)', vals.get("total_costs_incl_vat", vals.get("total_costs", 0)), CURRENCY_FMT)
    row += 1

    # === Section 2: מצב יוצא ===
    row = _write_section_header(ws, row, "מצב יוצא (Proposed State)")
    row = _write_kv_row(ws, row, 'סה"כ יחידות חדשות', int(vals.get("total_new_units", vals.get("total_units", 0))))
    row = _write_kv_row(ws, row, 'שטחי תמורות (מ"ר)', vals.get("total_return_floorplate", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, 'סה"כ שטח (מ"ר)', vals.get("total_floorplate", vals.get("total_residential_area", 0)), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "יחידות יזם", int(vals.get("developer_units", 0)))
    row = _write_kv_row(ws, row, 'שטח יזם (מ"ר)', vals.get("developer_floorplate", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, 'גודל ממוצע דירת יזם (מ"ר)', vals.get("avg_developer_unit_size", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "יחס שילוב", vals.get("combination_ratio", 0), DECIMAL_FMT)
    row += 1

    # === Section 3: פרוגרמה ===
    row = _write_section_header(ws, row, "פרוגרמה (Building Program)")
    row = _write_kv_row(ws, row, 'שטחי שירות (מ"ר)', vals.get("service_areas", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, 'סה"כ מעל קרקע (מ"ר)', vals.get("total_above_ground", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, 'שטח קומה (מ"ר)', vals.get("floor_area", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "מספר בניינים מקסימלי", int(vals.get("max_buildings", 0)))
    row = _write_kv_row(ws, row, 'מעל קרקע לבניין (מ"ר)', vals.get("above_ground_per_building", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, 'שטח פיתוח (מ"ר)', vals.get("development_land", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, 'מגורים לבניין (מ"ר)', vals.get("residential_per_building", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "יח' החזר לבניין", vals.get("return_units_per_building", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "יח' יזם לבניין", vals.get("developer_units_per_building", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, 'שטח יזם לבניין (מ"ר)', vals.get("developer_floorplate_per_building", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "חניות", int(vals.get("total_parking_spots", 0)))
    row = _write_kv_row(ws, row, 'שטח חניה (מ"ר)', vals.get("total_parking_area", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "קומות חניה", vals.get("parking_floors", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, 'שטח מרפסות (מ"ר)', vals.get("total_balcony_area", 0), DECIMAL_FMT)
    row += 1

    # === Apartment Mix ===
    if sim.apartment_mix:
        row = _write_section_header(ws, row, "תמהיל דירות")
        mix_rows = []
        for a in sim.apartment_mix:
            mix_rows.append([a.apartment_type, int(a.quantity), float(a.percentage_of_mix)])
        mix_rows.append(['סה"כ', int(vals.get("total_units", 0)), 100.0])
        row = _write_table(ws, row, ["סוג דירה", "כמות", "אחוז (%)"], mix_rows, [None, "#,##0", "0.0"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Economic Report (דוח כלכלי) — Sections 4-5
# ---------------------------------------------------------------------------

def generate_economic_report(sim: Simulation) -> io.BytesIO:
    """Generate detailed economic Excel report.

    Sections:
      4. עלויות (all cost items with breakdown)
      5. הכנסות + רווח (revenue, profit, IRR, NPV)
      + Cash flow table
      + Sensitivity matrix
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "דוח כלכלי"
    ws.sheet_view.rightToLeft = True

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    vals = _get_all_results(sim)
    cost_bd = vals.get("cost_breakdown", {})
    rev_bd = vals.get("revenue_breakdown", {})
    details = vals.get("calculation_details", {})

    project = sim.project if hasattr(sim, "project") else None

    # === Title ===
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value="דוח כלכלי מפורט - בדיקת כדאיות")
    title_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER_ALIGN
    row += 1
    ws.cell(row=row, column=1, value=f"פרויקט: {project.name if project else sim.version_name}").font = BOLD_FONT
    ws.cell(row=row, column=3, value=f"תאריך: {datetime.now().strftime('%d/%m/%Y')}").font = CELL_FONT
    row += 2

    # === Section 4: עלויות — Construction Costs ===
    row = _write_section_header(ws, row, "עלויות בנייה", 6)

    construction_items = [
        ("בנייה מגורים", "construction_residential"),
        ("בנייה שירות", "construction_service"),
        ("בנייה ציבורי/מסחרי", "construction_public"),
        ("מרפסות", "construction_balcony"),
        ("פיתוח", "construction_development"),
        ("חניה", "parking_construction"),
    ]
    for label, key in construction_items:
        row = _write_kv_row(ws, row, label, _safe(cost_bd.get(key, 0)), CURRENCY_FMT)
    row = _write_kv_row(ws, row, 'סה"כ בנייה', vals.get("construction_cost", 0), CURRENCY_FMT, bold=True)
    row += 1

    # === Additional Costs ===
    row = _write_section_header(ws, row, "עלויות נוספות", 6)

    additional_items = [
        ("היטל השבחה", "betterment_levy"),
        ("מס רכישה", "purchase_tax"),
        ("תכנון ויועצים", "planning_consultants"),
        ("היתרים ואגרות", "permits_fees"),
        ("חיבור חשמל", "electricity_connection"),
        ("פיקוח בנק", "bank_supervision"),
        ("ניהול הנדסי", "engineering_management"),
        ("פיקוח דיירים", "tenant_supervision"),
        ("הנהלה וכלליות", "management_overhead"),
        ("שיווק ופרסום", "marketing_advertising"),
        ("עו״ד דיירים", "tenant_lawyer"),
        ("דמי ייזום", "initiation_fee"),
        ("סבסוד שכ״ד", "rent_subsidy"),
        ("עלות פינוי", "evacuation_cost"),
        ("הובלות", "moving_cost"),
        ("בלת\"מ", "contingency"),
        ("עו״ד יזם", "developer_lawyer"),
        ("הריסה", "demolition"),
    ]
    for label, key in additional_items:
        val = _safe(cost_bd.get(key, 0))
        if val > 0:
            row = _write_kv_row(ws, row, label, val, CURRENCY_FMT)
    row += 1

    # === Financial Costs ===
    row = _write_section_header(ws, row, "עלויות מימון ומיסים", 6)
    row = _write_kv_row(ws, row, "עלות מימון", vals.get("financing_cost", 0), CURRENCY_FMT)
    # CPI linkage (audit gap)
    cp = sim.cost_parameters
    if cp and _safe(getattr(cp, "cpi_linkage_pct", None)) > 0:
        row = _write_kv_row(ws, row, "הצמדה למדד (%)", _safe(cp.cpi_linkage_pct), DECIMAL_FMT)
    row = _write_kv_row(ws, row, 'סה"כ עלויות ללא מע"מ', vals.get("total_costs_excl_vat", 0), CURRENCY_FMT, bold=True)
    row = _write_kv_row(ws, row, 'סה"כ עלויות כולל מע"מ', vals.get("total_costs_incl_vat", vals.get("total_costs", 0)), CURRENCY_FMT, bold=True, highlight="red")
    row += 1

    # === Section 5: הכנסות ===
    row = _write_section_header(ws, row, "הכנסות", 6)
    row = _write_kv_row(ws, row, "הכנסות מגורים", _safe(rev_bd.get("residential", vals.get("residential_revenue", 0))), CURRENCY_FMT)
    row = _write_kv_row(ws, row, "הכנסות מסחרי", _safe(rev_bd.get("commercial", vals.get("commercial_revenue", 0))), CURRENCY_FMT)
    # Parking & storage revenue (audit gap)
    parking_rev = _safe(rev_bd.get("parking", 0))
    if parking_rev > 0:
        row = _write_kv_row(ws, row, "הכנסות חניה", parking_rev, CURRENCY_FMT)
    storage_rev = _safe(rev_bd.get("storage", 0))
    if storage_rev > 0:
        row = _write_kv_row(ws, row, "הכנסות מחסנים", storage_rev, CURRENCY_FMT)
    marketing = _safe(rev_bd.get("marketing_discount", 0))
    if marketing > 0:
        row = _write_kv_row(ws, row, "הנחה שיווקית", -marketing, CURRENCY_FMT)
    row = _write_kv_row(ws, row, 'סה"כ הכנסות נטו', vals.get("net_revenue", vals.get("total_revenue", 0)), CURRENCY_FMT, bold=True, highlight="green")
    row += 1

    # === Profit ===
    row = _write_section_header(ws, row, "רווח", 6)
    profit = vals.get("expected_profit", vals.get("profit", 0))
    profit_hl = "green" if profit > 0 else "red"
    row = _write_kv_row(ws, row, "רווח יזמי (₪)", profit, CURRENCY_FMT, bold=True, highlight=profit_hl)
    row = _write_kv_row(ws, row, "שיעור רווחיות (%)", vals.get("profit_percent", vals.get("profitability_rate", 0)), DECIMAL_FMT, bold=True)
    row = _write_kv_row(ws, row, "רווחיות סטנדרט 21% (%)", vals.get("profit_percent_standard21", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "IRR (%)", vals.get("irr", 0), DECIMAL_FMT)
    row = _write_kv_row(ws, row, "NPV (₪)", vals.get("npv", 0), CURRENCY_FMT)
    row += 1

    # === Cash Flow Table ===
    cash_flows = vals.get("monthly_cash_flows", [])
    if cash_flows and isinstance(cash_flows, list):
        row = _write_section_header(ws, row, "תזרים מזומנים חודשי", 6)
        cf_rows = []
        cumulative = 0.0
        for i, cf in enumerate(cash_flows):
            cf_val = _safe(cf)
            cumulative += cf_val
            cf_rows.append([i + 1, cf_val, cumulative])
        row = _write_table(ws, row, ["חודש", "תזרים (₪)", "מצטבר (₪)"],
                          cf_rows, ["#,##0", CURRENCY_FMT, CURRENCY_FMT])

    # === Sensitivity Matrix ===
    row = _write_section_header(ws, row, "ניתוח רגישות - רווח יזמי", 6)

    base_revenue = _safe(vals.get("net_revenue", vals.get("total_revenue", 0)))
    base_costs = _safe(vals.get("total_costs_incl_vat", vals.get("total_costs", 0)))
    pct_changes = [-20, -10, 0, 10, 20]

    sens_headers = ["הכנסות \\ עלויות"] + [f"{p:+d}% עלויות" for p in pct_changes]
    for i, h in enumerate(sens_headers):
        ws.cell(row=row, column=i + 1, value=h)
    _style_row(ws, row, len(sens_headers), fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row += 1

    for rev_pct in pct_changes:
        adj_rev = base_revenue * (1 + rev_pct / 100)
        row_data = [f"{rev_pct:+d}% הכנסות"]
        for cost_pct in pct_changes:
            adj_cost = base_costs * (1 + cost_pct / 100)
            sensitivity_profit = adj_rev - adj_cost
            row_data.append(round(sensitivity_profit, 0))
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=row, column=c_idx + 1, value=val)
            if c_idx > 0:
                cell.number_format = CURRENCY_FMT
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = RED_FONT
                elif isinstance(val, (int, float)) and val > 0:
                    cell.font = GREEN_FONT
            # Highlight base case (0%/0%)
            if rev_pct == 0 and c_idx == 3:
                cell.fill = YELLOW_FILL
        _style_row(ws, row, len(sens_headers))
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
